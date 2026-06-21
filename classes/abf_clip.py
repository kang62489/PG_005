## Modules
# Standard library imports
from pathlib import Path

# Third-party imports
import numpy as np
import polars as pl
import pyabf
import tifffile
from openpyxl import Workbook
from rich.console import Console
from scipy.signal import find_peaks
from tabulate import tabulate

console = Console()


class AbfClip:
    def __init__(
        self,
        proc_tiff_path: Path,
        raw_abf_path: Path,
        results_dir: Path,
        detrend_mode: str,
        normalization: str,
        fs_imgs: float = 20,
    ) -> None:
        self.proc_tiff_path = proc_tiff_path
        self.raw_abf_path = raw_abf_path
        self.results_dir = results_dir
        self.detrend_mode = detrend_mode
        self.normalization = normalization
        self.fs_imgs = fs_imgs
        self.set_interval_frames: int = 0

        # Derive metadata from path names for export compatibility
        abf_parts = raw_abf_path.stem.split("_")
        self.exp_date = "_".join(abf_parts[:3])
        self.abf_serial = abf_parts[-1]
        self.img_serial = proc_tiff_path.stem.split("-")[1].split("_")[0]

        self.lst_img_frame_ranges: list[tuple[int, int]] = []
        self.lst_abf_sample_ranges: list[tuple[int, int]] = []
        self.df_skipped_spikes: pl.DataFrame | None = None
        self.df_picked_spikes: pl.DataFrame | None = None

        # Hardware constants
        self.TTL_5V_HIGH: float = 2.0
        self.TTL_5V_LOW: float = 0.8

        results_dir.mkdir(parents=True, exist_ok=True)

        with tifffile.TiffFile(proc_tiff_path) as tif:
            self.n_frames = len(tif.pages)

        self.load_abf()
        self.spike_detection()
        self.get_available_spiking_frames()
        self.clip_time_abf_img_segments()
        self._export_spike_xlsx()

    def load_abf(self) -> None:
        self.loaded_abf = pyabf.ABF(self.raw_abf_path)

    def spike_detection(
        self, spike_min_distance: int = 3000, spike_min_prominence: float = 20.0
    ) -> None:
        self.abf_time = self.loaded_abf.sweepX
        self.abf_dataset = self.loaded_abf.data
        self.abf_fs = self.loaded_abf.dataRate

        self.abf_idx_tstart: int = np.where(self.abf_dataset[3] >= self.TTL_5V_HIGH)[0][0]
        self.abf_idx_tend: int = (
            len(self.abf_dataset[3]) - np.where(np.flip(self.abf_dataset[3]) >= self.TTL_5V_LOW)[0][0]
        )

        self.Vm: np.ndarray = self.abf_dataset[0][self.abf_idx_tstart : self.abf_idx_tend]
        self.rec_time: np.ndarray = self.abf_time[self.abf_idx_tstart : self.abf_idx_tend]

        self.peak_indices, _properties = find_peaks(
            self.Vm, distance=spike_min_distance, prominence=spike_min_prominence
        )
        self.num_found_spikes = len(self.peak_indices)

        console.log(f"Membrane Potential Sampling Rate: {self.abf_fs} Hz")
        console.log(f"Start index: {self.abf_idx_tstart}, Start time: {self.abf_time[self.abf_idx_tstart]}")
        console.log(f"End index: {self.abf_idx_tend}, End time: {self.abf_time[self.abf_idx_tend]}")
        console.log(f"Found {self.num_found_spikes} peaks")

        self.df_Vm = pl.DataFrame({"Time": self.rec_time, "Vm": self.Vm})
        self.peak_times = self.rec_time[self.peak_indices]
        self.peak_values = self.Vm[self.peak_indices]
        self.df_peaks = pl.DataFrame({"Time": self.peak_times, "Peaks": self.peak_values})

    def _export_spike_xlsx(self) -> None:
        wb = Workbook()

        ws_vm = wb.active
        ws_vm.title = "Vm"
        ws_vm.append(self.df_Vm.columns)
        for row in self.df_Vm.iter_rows(named=False):
            ws_vm.append(list(row))

        ws_peaks = wb.create_sheet("Peaks")
        ws_peaks.append(self.df_peaks.columns)
        for row in self.df_peaks.iter_rows(named=False):
            ws_peaks.append(list(row))

        if self.lst_abf_sample_ranges:
            ws_segs = wb.create_sheet("ABF_segments")
            headers: list = []
            for i in range(len(self.lst_abf_sample_ranges)):
                headers += [f"rec_time(abf_seg_{i})", f"Vm(abf_seg_{i})"]
            ws_segs.append(headers)

            segments = [
                (self.rec_time[start:end], self.Vm[start:end])
                for start, end in self.lst_abf_sample_ranges
            ]
            for r in range(len(segments[0][0])):
                row: list = []
                for time_seg, vm_seg in segments:
                    row += [float(time_seg[r]), float(vm_seg[r])]
                ws_segs.append(row)

        xlsx_dir = self.results_dir / "xlsx"
        xlsx_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = xlsx_dir / f"ABF_{self.exp_date}_{self.abf_serial}_spike_analysis.xlsx"
        wb.save(xlsx_path)
        console.log(f"[green]Saved spike detection → {xlsx_path}[/green]")

    def get_available_spiking_frames(self) -> None:
        if self.num_found_spikes == 0:
            console.log("[bold red]No peak is found! Exiting...[/bold red]")
            return

        self.ts_imgs: float = 1 / self.fs_imgs
        self.points_per_frame: int = int(self.ts_imgs * self.abf_fs)

        self.total_triggered_frames = (self.abf_idx_tend - self.abf_idx_tstart) // self.points_per_frame
        console.log(f"Total recorded frames: {self.total_triggered_frames}")

        self.abf_img_frame_num_diff = self.total_triggered_frames - self.n_frames
        if self.abf_img_frame_num_diff == 0:
            self.first_dropped = 0
            self.last_dropped = 0
        else:
            self.first_dropped = 1
            self.last_dropped = self.abf_img_frame_num_diff - 1

        self.spike_frame_indices = np.floor(self.peak_indices / self.points_per_frame).astype(int) - self.first_dropped

        inter_spike_frames = (np.diff(self.spike_frame_indices) - 1).astype(int)
        leading_interval_frames: int = self.spike_frame_indices[0] - 1
        trailing_inverval_frames: int = self.n_frames - self.spike_frame_indices[-1] - 1
        inter_spike_frames = np.insert(inter_spike_frames, 0, leading_interval_frames).astype(int)
        inter_spike_frames = np.append(inter_spike_frames, trailing_inverval_frames).astype(int)
        # Spikes closer together than 1 frame (after flooring to frame index) can make this
        # negative, which would later flip set_interval_frames negative and invert left/right
        # bounds into an empty frame range — clamp to 0 ("no margin available") instead.
        inter_spike_frames = np.clip(inter_spike_frames, 0, None)

        # Pass 1: collect min_available_frames for all spikes → derive set_interval_frames from mode
        all_min_available: list[int] = []
        for idx_of_spike in range(len(self.spike_frame_indices)):
            left_frames: int = inter_spike_frames[idx_of_spike]
            right_frames: int = inter_spike_frames[idx_of_spike + 1]
            all_min_available.append(int(np.min([left_frames, right_frames])))

        all_min_series = pl.Series("Min_Available_Frames", all_min_available, dtype=pl.Int64)
        self.set_interval_frames = min(int(all_min_series.mode().min()), 20)
        console.log(
            f"[bold cyan]Min_Available_Frames — max: {all_min_series.max()}, "
            f"mode: {all_min_series.mode().to_list()} → set_interval_frames = {self.set_interval_frames}[/bold cyan]"
        )

        # Pass 2: filter spikes using auto-derived set_interval_frames
        lst_skipped_spikes = []
        lst_picked_spikes = []

        for idx_of_spike, frame_of_spike in enumerate(self.spike_frame_indices):
            min_available_frames = all_min_available[idx_of_spike]

            # set_interval_frames == 0 means every segment would be just the spike frame itself,
            # with no baseline frames before it — unanalyzable downstream (zscore_img_segs/
            # SpatialCategorizer both need at least 1 baseline frame). Skip rather than crash.
            if self.set_interval_frames < 1 or min_available_frames < self.set_interval_frames:
                lst_skipped_spikes.append(
                    {"Spike_Frame_Index": frame_of_spike, "Min_Available_Frames": min_available_frames}
                )
                continue

            lst_picked_spikes.append(
                {
                    "Spike_Frame_Index": frame_of_spike,
                    "Min_Available_Frames": min_available_frames,
                    "Set_Interval_Frames": self.set_interval_frames,
                }
            )

        self.df_skipped_spikes = pl.DataFrame(
            lst_skipped_spikes,
            schema={"Spike_Frame_Index": pl.Int64, "Min_Available_Frames": pl.Int64},
        )
        self.df_picked_spikes = pl.DataFrame(
            lst_picked_spikes,
            schema={"Spike_Frame_Index": pl.Int64, "Min_Available_Frames": pl.Int64, "Set_Interval_Frames": pl.Int64},
        )

        console.log("\n[bold red]" + "-" * 32 + " Skipped Spikes " + "-" * 32 + "\n[/bold red]")
        console.log(
            f"[bold red]Total skipped spikes: {len(self.df_skipped_spikes)}/{self.num_found_spikes}[/bold red]"
        )
        console.log(tabulate(self.df_skipped_spikes.to_dicts(), headers="keys", showindex=False, tablefmt="pretty"))
        console.log("[bold red]" + "=" * 80 + "\n" + "[/bold red]")

        console.log("\n[bold green]" + "-" * 32 + " Picked Spikes " + "-" * 32 + "\n[/bold green]")
        console.log(
            f"[bold green]Total picked spikes: {len(self.df_picked_spikes)}/{self.num_found_spikes}[/bold green]"
        )
        console.log(tabulate(self.df_picked_spikes.to_dicts(), headers="keys", showindex=False, tablefmt="pretty"))
        console.log("[bold green]" + "=" * 79 + "\n" + "[/bold green]")

    def clip_time_abf_img_segments(self) -> None:
        if self.df_picked_spikes is None:
            console.log("[bold red]No spikes were picked! Exiting...[/bold red]")
            return

        for row in self.df_picked_spikes.iter_rows(named=True):
            left_bound: int = row["Spike_Frame_Index"] - row["Set_Interval_Frames"]
            right_bound: int = row["Spike_Frame_Index"] + row["Set_Interval_Frames"]

            abf_left_bound: int = left_bound + self.first_dropped
            abf_right_bound: int = right_bound + self.first_dropped

            self.lst_img_frame_ranges.append((left_bound, right_bound))
            start_sample = abf_left_bound * self.points_per_frame
            end_sample = (abf_right_bound + 1) * self.points_per_frame
            self.lst_abf_sample_ranges.append((start_sample, end_sample))

        console.log(
            f"Total segments: {len(self.lst_img_frame_ranges)} image ranges, {len(self.lst_abf_sample_ranges)} ABF ranges"
        )

    # ── Export ─────────────────────────────────────────────────────────────────

    def get_export_data(self) -> dict:
        return {
            "exp_date": self.exp_date,
            "tiff_full_path": self.proc_tiff_path,
            "abf_full_path": self.raw_abf_path,
            "abf_serial": self.abf_serial,
            "img_serial": self.img_serial,
            "num_found_spikes": self.num_found_spikes,
            "n_spikes_analyzed": len(self.df_picked_spikes),
        }
