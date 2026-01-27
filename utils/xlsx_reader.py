"""Read metadata from REC_*.xlsx files in rec_summary directory."""

from pathlib import Path

from openpyxl import load_workbook


def get_picked_pairs(rec_summary_dir: str | Path | None = None) -> list[dict]:
    """
    Read all REC_*.xlsx files and return list of picked pairs.

    Args:
        rec_summary_dir: Directory containing REC_*.xlsx files
                         If None, uses path from config_paths

    Returns:
        list of dict: [{'exp_date': '2025_12_18', 'img_serial': '0026',
                        'abf_serial': '0023', 'objective': '10X'}, ...]
    """
    # Default to config path
    if rec_summary_dir is None:
        from config_paths import PATHS
        rec_summary_dir = PATHS["rec_summary"]
    else:
        rec_summary_dir = Path(rec_summary_dir)

    pairs = []

    for xlsx_file in rec_summary_dir.glob("REC_*.xlsx"):
        # Extract date from filename: REC_2025_12_18.xlsx → 2025_12_18
        exp_date = xlsx_file.stem.replace("REC_", "")

        wb = load_workbook(xlsx_file)
        ws = wb.active

        # Get column indices
        headers = [cell.value for cell in ws[1]]

        # Find ABF column (different files use different names)
        abf_col_names = ["ABF_NUMBER", "ABF_SERIAL_NUMBER", "ABF_SERIAL", "ABF", "STIM_PROTOCOL_NUMBER"]
        col_abf = None
        for col_name in abf_col_names:
            if col_name in headers:
                col_abf = headers.index(col_name)
                break

        # Skip this file if no ABF column or PICK column found
        if col_abf is None or "PICK" not in headers or "Filename" not in headers or "OBJ" not in headers:
            continue

        col_filename = headers.index("Filename")
        col_obj = headers.index("OBJ")
        col_pick = headers.index("PICK")

        # Read rows where PICK is not empty
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[col_pick]:  # If PICK column has value
                filename = row[col_filename]  # "2025_12_18-0026.tif"
                if not filename:
                    continue

                img_serial = filename.split("-")[1].replace(".tif", "")  # "0026"
                abf_serial = row[col_abf]  # "0023"
                objective = row[col_obj]  # "10X"

                # Skip if any required field is missing
                if not abf_serial or not objective:
                    continue

                pairs.append(
                    {"exp_date": exp_date, "img_serial": img_serial, "abf_serial": abf_serial, "objective": objective}
                )

    return pairs
